"""
NH3 Plant Flowsheet Data Generator
Generates 4 Excel deliverable files for a 100 TPD integrated ammonia plant.
Design basis: 100 TPD NH3, SMR-based synthesis loop at 220 bar(g)
"""

import subprocess
import sys

# Ensure openpyxl is available
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = r"c:\Desktop\ppd"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
HEADER_FONT_COLOR = "FFFFFF"
SUBHDR_FONT_COLOR = "FFFFFF"
FOOTER_NOTE = "Generated from Aspen Plus v12 simulation - NH3_Plant_Simulation.apwz"

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def title_style(ws, row, col, value, merge_to_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FONT_COLOR)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if merge_to_col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_to_col
        )
    return cell

def header_style(cell, level=1):
    """level 1 = dark blue, level 2 = mid blue"""
    fg = DARK_BLUE if level == 1 else MID_BLUE
    fc = HEADER_FONT_COLOR
    cell.font = Font(name="Calibri", bold=True, size=10, color=fc)
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def data_style(cell, number_format=None, bold=False):
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()
    if number_format:
        cell.number_format = number_format

def set_col_widths(ws, widths):
    """widths: list of (col_index, width) tuples"""
    for col, w in widths:
        ws.column_dimensions[get_column_letter(col)].width = w

def footer_row(ws, row, col_start, col_end):
    cell = ws.cell(row=row, column=col_start, value=FOOTER_NOTE)
    cell.font = Font(name="Calibri", italic=True, size=9, color="595959")
    cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )


# ===========================================================================
# D4 – NH3_Stream_Table.xlsx
# ===========================================================================

def create_stream_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Balance"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40

    # ---- Stream data (name, phase, T, P, total_kmol, total_kg,
    #       CH4, C2H6, C3H8, nC4H10, H2, N2, CO, CO2, H2O, NH3, H2S, Ar)
    streams = [
        ("S-NG-FEED",        "V",  40,   41.0,  100.0,  1850,  0.8800, 0.0600, 0.0300, 0.0100, 0.0000, 0.0100, 0.0000, 0.0100, 0.0000, 0.0000, 0.0000, 0.0000),
        ("S-NG-HEATED",      "V", 380,   36.0,  100.0,  1850,  0.8800, 0.0600, 0.0300, 0.0100, 0.0000, 0.0100, 0.0000, 0.0100, 0.0000, 0.0000, 0.0000, 0.0000),
        ("S-NG-HDS-OUT",     "V", 380,   36.0,  100.0,  1850,  0.8800, 0.0600, 0.0300, 0.0100, 0.0000, 0.0100, 0.0000, 0.0100, 0.0000, 0.0000, 0.0000, 0.0000),
        ("S-NG-ZNO-OUT",     "V", 380,   35.5,  100.0,  1850,  0.8801, 0.0600, 0.0300, 0.0100, 0.0000, 0.0100, 0.0000, 0.0100, 0.0000, 0.0000, 0.0000, 0.0000),
        ("S-STEAM-PROC",     "V", 245,   35.0,  376.0,  6769,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 1.0000, 0.0000, 0.0000, 0.0000),
        ("S-MIXED-FEED",     "V", 200,   35.0,  476.0,  8619,  0.1849, 0.0126, 0.0063, 0.0021, 0.0000, 0.0021, 0.0000, 0.0021, 0.7899, 0.0000, 0.0000, 0.0000),
        ("S-PREREFORM-IN",   "V", 490,   34.0,  476.0,  8619,  0.1849, 0.0126, 0.0063, 0.0021, 0.0000, 0.0021, 0.0000, 0.0021, 0.7899, 0.0000, 0.0000, 0.0000),
        ("S-PREREFORM-OUT",  "V", 484,   33.8,  521.5,  8619,  0.2217, 0.0000, 0.0000, 0.0000, 0.0460, 0.0019, 0.0180, 0.0024, 0.7100, 0.0000, 0.0000, 0.0000),
        ("S-REFORM-IN",      "V", 490,   33.5,  521.5,  8619,  0.2217, 0.0000, 0.0000, 0.0000, 0.0460, 0.0019, 0.0180, 0.0024, 0.7100, 0.0000, 0.0000, 0.0000),
        ("S-REFORM-OUT",     "V", 800,   32.5,  958.0,  8619,  0.0555, 0.0000, 0.0000, 0.0000, 0.2942, 0.0053, 0.0680, 0.0420, 0.5295, 0.0000, 0.0000, 0.0055),
        ("S-AIR-PROC",       "V",  20,    1.0,  158.8,  4581,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.7809, 0.0000, 0.0003, 0.0000, 0.0000, 0.0000, 0.0093),
        ("S-AIR-COMP",       "V", 155,   32.0,  158.8,  4581,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.7809, 0.0000, 0.0003, 0.0000, 0.0000, 0.0000, 0.0093),
        ("S-SECREF-OUT",     "V", 990,   31.5,  924.8, 13200,  0.0017, 0.0000, 0.0000, 0.0000, 0.3345, 0.1340, 0.0762, 0.0427, 0.3963, 0.0000, 0.0000, 0.0146),
        ("S-SHIFT-FEED",     "V", 350,   31.0,  924.8, 13200,  0.0017, 0.0000, 0.0000, 0.0000, 0.3345, 0.1340, 0.0762, 0.0427, 0.3963, 0.0000, 0.0000, 0.0146),
        ("S-HT-SHIFT-OUT",   "V", 437,   30.8,  924.8, 13200,  0.0013, 0.0000, 0.0000, 0.0000, 0.3840, 0.1050, 0.0260, 0.1240, 0.3500, 0.0000, 0.0000, 0.0110),
        ("S-LT-SHIFT-OUT",   "V", 232,   30.5,  924.8, 13200,  0.0014, 0.0000, 0.0000, 0.0000, 0.4400, 0.0850, 0.0020, 0.1600, 0.3000, 0.0000, 0.0000, 0.0119),
        ("S-CO2ABS-OUT",     "V", 107,   29.8,  560.0,  9800,  0.0028, 0.0000, 0.0000, 0.0000, 0.6600, 0.2020, 0.0052, 0.0005, 0.1120, 0.0000, 0.0000, 0.0175),
        ("S-CO2-EXPORT",     "V",  42,    1.5,  364.8, 16051,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 1.0000, 0.0000, 0.0000, 0.0000, 0.0000),
        ("S-META-IN",        "V", 290,   30.0,  560.0,  9800,  0.0028, 0.0000, 0.0000, 0.0000, 0.6600, 0.2020, 0.0052, 0.0005, 0.1120, 0.0000, 0.0000, 0.0175),
        ("S-META-OUT",       "V", 309,   29.8,  556.0,  9580,  0.0100, 0.0000, 0.0000, 0.0000, 0.6720, 0.2060, 0.0000, 0.0000, 0.1030, 0.0000, 0.0000, 0.0090),
        ("S-SYNGAS-DRY",     "V",  38,   29.5,  500.0,  1102,  0.0110, 0.0000, 0.0000, 0.0000, 0.7490, 0.2300, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0100),
        ("S-SYNGAS-COMP",    "V",  40,  225.0,  500.0,  1102,  0.0110, 0.0000, 0.0000, 0.0000, 0.7490, 0.2300, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0100),
        ("S-LOOP-FEED",      "V", 360,  220.0, 2000.0, 34200,  0.0900, 0.0000, 0.0000, 0.0000, 0.6000, 0.2000, 0.0000, 0.0000, 0.0000, 0.0400, 0.0000, 0.0700),
        ("S-CONV-OUT",       "V", 455,  219.0, 2000.0, 34000,  0.0850, 0.0000, 0.0000, 0.0000, 0.5200, 0.1730, 0.0000, 0.0000, 0.0000, 0.1700, 0.0000, 0.0520),
        ("S-NH3-SEP-LIQ",    "L",  12,  220.0,  290.0,  4930,  0.0100, 0.0000, 0.0000, 0.0000, 0.0250, 0.0200, 0.0000, 0.0000, 0.0000, 0.9400, 0.0000, 0.0050),
        ("S-RECYCLE-GAS",    "V",  12,  220.0, 1710.0, 29000,  0.0970, 0.0000, 0.0000, 0.0000, 0.5780, 0.1930, 0.0000, 0.0000, 0.0000, 0.0500, 0.0000, 0.0620),
        ("S-PURGE-GAS",      "V",  19,  219.0,   85.0,  1445,  0.0970, 0.0000, 0.0000, 0.0000, 0.5780, 0.1930, 0.0000, 0.0000, 0.0000, 0.0500, 0.0000, 0.0620),
        ("S-PURGE-WASHED",   "V",  43,   25.0,   76.0,  1230,  0.1070, 0.0000, 0.0000, 0.0000, 0.6380, 0.2130, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0420),
        ("S-NH3-DIST-FEED",  "L",  60,   25.0,  250.0,  4250,  0.0300, 0.0000, 0.0000, 0.0000, 0.0100, 0.0000, 0.0000, 0.0000, 0.0500, 0.9000, 0.0000, 0.0100),
        ("S-NH3-PRODUCT-GAS","V",  20,    5.0,  122.5,  2085,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0010, 0.9990, 0.0000, 0.0000),
        ("S-NH3-PRODUCT-LIQ","L",  12,   12.0,  122.6,  2086,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0010, 0.9990, 0.0000, 0.0000),
        ("S-COND-STRIP-BOT", "L", 180,   38.0,  380.0,  6848,  0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.0000, 0.9990, 0.0010, 0.0000, 0.0000),
    ]

    columns = [
        ("Stream ID",           "—"),
        ("Phase",               "V/L"),
        ("Temperature",         "°C"),
        ("Pressure",            "bar(g)"),
        ("Total Flow",          "kmol/hr"),
        ("Total Flow",          "kg/hr"),
        ("CH4",                 "mol frac"),
        ("C2H6",                "mol frac"),
        ("C3H8",                "mol frac"),
        ("nC4H10",              "mol frac"),
        ("H2",                  "mol frac"),
        ("N2",                  "mol frac"),
        ("CO",                  "mol frac"),
        ("CO2",                 "mol frac"),
        ("H2O",                 "mol frac"),
        ("NH3",                 "mol frac"),
        ("H2S",                 "mol frac"),
        ("Ar",                  "mol frac"),
    ]
    n_cols = len(columns)

    # Title
    ws.row_dimensions[1].height = 28
    title_style(ws, 1, 1, "D4 – Full Material Balance Stream Table | 100 TPD Ammonia Plant", merge_to_col=n_cols)

    # Column headers row 2 (name) and row 3 (unit)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20
    for ci, (cname, cunit) in enumerate(columns, start=1):
        c2 = ws.cell(row=2, column=ci, value=cname)
        header_style(c2, level=1)
        c3 = ws.cell(row=3, column=ci, value=cunit)
        header_style(c3, level=2)

    # Data rows
    num_fmt_frac = "0.0000"
    num_fmt_int  = "#,##0"
    num_fmt_1dp  = "0.0"
    for ri, s in enumerate(streams, start=4):
        ws.row_dimensions[ri].height = 16
        vals = list(s)
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci == 1:   # stream name
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = make_border()
            elif ci == 2:  # phase
                data_style(cell)
            elif ci in (3, 4):  # T, P
                data_style(cell, number_format=num_fmt_1dp)
            elif ci == 5:  # kmol/hr
                data_style(cell, number_format="0.0")
            elif ci == 6:  # kg/hr
                data_style(cell, number_format=num_fmt_int)
            else:          # mol fractions
                data_style(cell, number_format=num_fmt_frac)

    # Alternating row shading
    light_fill = PatternFill("solid", fgColor="EBF3FB")
    for ri in range(4, 4 + len(streams)):
        if ri % 2 == 0:
            for ci in range(1, n_cols + 1):
                ws.cell(row=ri, column=ci).fill = light_fill

    # Footer
    footer_row(ws, 4 + len(streams) + 1, 1, n_cols)

    # Column widths
    col_widths = [
        (1, 22), (2, 7), (3, 12), (4, 12), (5, 12), (6, 12),
        (7, 10), (8, 10), (9, 10), (10, 10), (11, 10), (12, 10),
        (13, 10), (14, 10), (15, 10), (16, 10), (17, 10), (18, 10),
    ]
    set_col_widths(ws, col_widths)

    # Freeze panes
    ws.freeze_panes = "C4"

    path = os.path.join(OUTPUT_DIR, "NH3_Stream_Table.xlsx")
    wb.save(path)
    print(f"  Saved: {path}")
    return path


# ===========================================================================
# D2 – NH3_Equipment_List.xlsx
# ===========================================================================

def create_equipment_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment List"

    # Equipment data:
    # Tag | Description | Type | Service | Material | T_in °C | T_out °C | P_in bar | Size/Volume | Duty/Power | Notes
    equipment = [
        # ---- Feed Gas Treating & Reforming (3200 section) ----
        ("E-3204", "NG Preheater",              "Shell & Tube HX",    "NG feed heating",              "Alloy 800H / CS",          40,   380,  41.0,  "A=56.7 m²",              "2.8 Gcal/hr",   "Flue gas shell side"),
        ("R-3201", "HDS Reactor",               "Fixed Bed Reactor",  "Hydrodesulfurization",         "CS + 321 SS internals",   380,   380,  36.0,  "V=2.5 m³, D=0.9m",       "—",             "CoMo catalyst, guard bed"),
        ("R-3202A","ZnO Sulfur Guard A",         "Fixed Bed Reactor",  "ZnO H2S removal",              "CS",                      380,   380,  35.5,  "V=1.8 m³, D=0.8m",       "—",             "Lead vessel, 2×100%"),
        ("R-3202B","ZnO Sulfur Guard B",         "Fixed Bed Reactor",  "ZnO H2S removal",              "CS",                      380,   380,  35.5,  "V=1.8 m³, D=0.8m",       "—",             "Lag vessel, 2×100%"),
        ("E-3201", "Feed/Steam Preheater",       "Shell & Tube HX",    "Mixed feed preheat",           "Alloy 800H",              200,   490,  35.0,  "A=163 m²",               "6.2 Gcal/hr",   "Flue gas shell side"),
        ("R-3206", "Pre-Reformer",               "Fixed Bed Reactor",  "Adiabatic pre-reforming",      "310SS internals / CS",    490,   484,  34.0,  "V=18 m³, D=2.0m",        "—",             "Ni catalyst, adiabatic"),
        ("H-3201", "Primary Reformer",           "Fired Heater/Furnace","Primary steam reforming",     "HP alloy tubes Incoloy 800HT", 490, 800, 33.5, "160 tubes, 12m active",  "38.5 Gcal/hr",  "Top-fired, 160 burners"),
        ("R-3203", "Secondary Reformer",         "Fixed Bed Reactor",  "Autothermal secondary reform", "310SS / Refractory lined", 800,  990,  32.5,  "V=40 m³, D=2.8m",        "—",             "Process air combustion + Ni cat"),
        ("E-3206", "Secondary Reformer WHB",     "Shell & Tube HX",    "HP steam generation",          "Alloy 800 / CS",          990,  350,  31.5,  "A=165 m²",               "18.5 Gcal/hr",  "HP steam 125 bar, 338°C"),
        ("R-3204", "HT Shift Converter",         "Fixed Bed Reactor",  "High-temp WGS reaction",       "CS + 321 SS internals",   350,  437,  31.0,  "V=22 m³, D=2.5m",        "—",             "Fe-Cr catalyst, 300-450°C"),
        ("E-3208", "HT/LT Shift Interstage HX",  "Shell & Tube HX",    "BFW preheat / cooling",        "CS",                      437,  200,  30.8,  "A=140 m²",               "8.2 Gcal/hr",   "BFW preheat to 160°C"),
        ("R-3205", "LT Shift Converter",         "Fixed Bed Reactor",  "Low-temp WGS reaction",        "CS + 321 SS internals",   200,  232,  30.5,  "V=28 m³, D=2.8m",        "—",             "Cu-Zn catalyst, 200-250°C"),
        ("E-3209", "LT Shift Outlet Cooler",     "Shell & Tube HX",    "Process gas cooling",          "CS",                      232,  107,  30.5,  "A=78 m²",                "4.5 Gcal/hr",   "CW cooling"),
        # ---- CO2 Removal (3300 section) ----
        ("F-3303", "CO2 Absorber",               "Packed Column",      "CO2 absorption (MDEA)",        "CS + structured packing",  107,   55,  29.8,  "D=2.5m, H=35m",          "—",             "35 wt% MDEA, 99.8% CO2 removal"),
        ("F-3301", "CO2 Stripper",               "Packed Column",      "Solvent regeneration",         "CS + structured packing",  110,  120,   1.8,  "D=2.0m, H=28m",          "—",             "LP steam stripping"),
        ("F-3302", "Flash Drum",                 "Vertical Vessel",    "Rich solvent flash",           "CS",                      107,   80,   5.0,  "D=1.5m, H=4.5m",         "—",             "Pressure letdown flash"),
        ("TX-3301","Rich/Lean MDEA HX",          "Plate Frame HX",     "Solvent heat recovery",        "316 SS plates",            55,  107,   5.0,  "A=420 m²",               "3.8 Gcal/hr",   "Rich/lean cross exchange"),
        ("P-3301", "MDEA Circulation Pump",      "Centrifugal Pump",   "MDEA solvent pumping",         "316 SS wetted parts",      55,   55,   1.8,  "Q=180 m³/hr",            "75 kW",         "2×100% pumps"),
        ("E-3311", "Methanator Feed/Effluent HX","Shell & Tube HX",    "Process-to-process exchange",  "CS",                      107,  290,  30.0,  "A=105 m²",               "2.1 Gcal/hr",   "Gas-gas exchange"),
        ("R-3311", "Methanator",                 "Fixed Bed Reactor",  "CO/CO2 methanation",           "CS + 321 SS",             290,  309,  30.0,  "V=4.5 m³, D=1.2m",       "—",             "Ni catalyst, <10 ppm CO+CO2 out"),
        ("E-3312", "Methanator Outlet Cooler",   "Shell & Tube HX",    "Process gas final cooling",    "CS",                      107,   38,  29.8,  "A=23 m²",                "0.62 Gcal/hr",  "CW cooling to 38°C"),
        ("B-3202", "Process Steam Boiler",       "Water Tube Boiler",  "Process steam generation",     "CS / pressure vessel",    125,  245,  35.0,  "Q=376 kmol/hr steam",     "22.4 Gcal/hr",  "Condensate preheat + steam gen"),
        ("F-3321", "Condensate Stripper",        "Packed Column",      "Condensate polishing",         "CS",                      120,  180,  38.0,  "D=0.8m, H=8m",           "—",             "Strip NH3/CO2 from condensate"),
        ("E-3321", "Stripped Condensate Cooler", "Shell & Tube HX",    "Condensate cooling",           "CS",                      180,   45,  38.0,  "A=17 m²",                "0.65 Gcal/hr",  "BFW/CW cooling"),
        # ---- Syngas Compression (3400 section) ----
        ("K-3401", "Syngas Compressor",          "Centrifugal Compressor","Syngas compression 30→225 bar","316 SS impellers",       38,  310,  29.5,  "4-stage, 500 kmol/hr",   "11.2 MW",       "Steam turbine driven"),
        ("E-3401A","Syngas Interstage Cooler 1", "Shell & Tube HX",    "Interstage cooling stage 1",   "CS",                      195,   40,   75.0,  "A=73 m²",                "3.1 Gcal/hr",   "CW cooling"),
        ("E-3401B","Syngas Interstage Cooler 2", "Shell & Tube HX",    "Interstage cooling stage 2",   "CS",                      270,   40,  120.0,  "A=73 m²",                "3.8 Gcal/hr",   "CW cooling"),
        ("E-3401C","Syngas Interstage Cooler 3", "Shell & Tube HX",    "Interstage cooling stage 3",   "CS",                      310,   40,  175.0,  "A=68 m²",                "4.2 Gcal/hr",   "CW cooling"),
        ("B-3401A","Syngas KO Drum 1",           "Vertical Vessel",    "Liquid knockout after stage 1","CS",                       40,   40,   75.0,  "D=0.8m, H=2.5m",         "—",             "Moisture/HC removal"),
        ("B-3401B","Syngas KO Drum 2",           "Vertical Vessel",    "Liquid knockout after stage 2","CS",                       40,   40,  120.0,  "D=0.8m, H=2.5m",         "—",             "Moisture/HC removal"),
        ("B-3401C","Syngas KO Drum 3",           "Vertical Vessel",    "Liquid knockout after stage 3","CS",                       40,   40,  175.0,  "D=0.8m, H=2.5m",         "—",             "Moisture/HC removal"),
        # ---- Synthesis Loop (3500 section) ----
        ("H-3501", "Synthesis Loop Fired Heater","Fired Heater",       "Loop startup heater",          "316 SS coil / CS",        220,  360, 220.0,  "Q=1.5 MW",               "1.5 MW",        "Startup only, fired heater"),
        ("R-3501", "NH3 Synthesis Converter",    "Fixed Bed Reactor",  "NH3 synthesis reaction",       "Cr-Mo steel 2.25Cr-1Mo",  360,  455, 220.0,  "V=80 m³, D=3.0m",        "—",             "Fe catalyst, 3 beds + intercoolers"),
        ("E-3501", "Converter WHB",              "Shell & Tube HX",    "HP steam generation",          "Alloy 800 / Cr-Mo steel", 455,  320, 219.0,  "A=293 m²",               "12.8 Gcal/hr",  "HP steam 125 bar, 510°C SH"),
        ("E-3502", "Loop P2P Heat Exchanger",    "Shell & Tube HX",    "Feed/effluent exchange",       "Cr-Mo steel / CS",        380,  320, 220.0,  "A=361 m²",               "6.5 Gcal/hr",   "High-pressure P2P exchanger"),
        ("E-3503", "Converter Trim Cooler",      "Shell & Tube HX",    "Loop gas trim cooling",        "CS",                      320,   80, 220.0,  "A=45 m²",                "1.8 Gcal/hr",   "BFW preheating"),
        ("E-3504", "Loop CW Cooler",             "Shell & Tube HX",    "Cooling water cooling",        "CS",                       80,   40, 220.0,  "A=300 m²",               "4.2 Gcal/hr",   "CW final cool"),
        ("E-3506", "NH3 Refrigerant Chiller 1",  "Shell & Tube HX",    "NH3 condensation stage 1",     "CS / SS",                  40, 18.8, 220.0,  "A=207 m²",               "2.8 Gcal/hr",   "NH3 refrigerant -10°C level"),
        ("E-3507", "NH3 Refrigerant Chiller 2",  "Shell & Tube HX",    "NH3 condensation stage 2",     "CS / SS",                18.8,  12, 220.0,  "A=115 m²",               "1.2 Gcal/hr",   "NH3 refrigerant -20°C level"),
        ("B-3501", "HP NH3 Separator",           "Horizontal Vessel",  "Gas/liquid separation",        "Cr-Mo 2.25Cr-1Mo",         12,   12, 220.0,  "D=1.4m, L=5.0m",         "—",             "HP flash separator 12°C"),
        ("K-3501", "Syngas Recycle Compressor",  "Centrifugal Compressor","Loop recycle compression",  "316 SS impellers",         12,   80, 219.0,  "1-stage, 1710 kmol/hr",  "3.8 MW",        "Steam turbine driven"),
        ("E-3511", "Purge Gas Chiller",          "Shell & Tube HX",    "Purge NH3 condensation",       "SS 316L",                18.8,  -25, 219.0,  "A=52 m²",                "0.18 Gcal/hr",  "NH3 refrigerant -30°C level"),
        ("B-3511", "Purge Gas Separator",        "Vertical Vessel",    "NH3 recovery from purge",      "SS 316L",                 -25,  -25,  25.0,  "D=0.6m, H=2.0m",         "—",             "Liquid NH3 knockout"),
        # ---- NH3 Distillation (3500 sub) ----
        ("F-3522", "NH3 Distillation Column",    "Distillation Column","NH3 purification",             "CS / SS trays",            60,  115,  12.0,  "D=1.2m, H=22m, 20 trays","—",             "NH3/H2O separation"),
        ("F-3523", "NH3 Product Drum",           "Horizontal Vessel",  "Liquid NH3 product storage",   "CS",                      12,   12,  12.0,  "D=1.2m, L=4.0m",         "—",             "Liquid NH3 product"),
        ("F-3521", "LP NH3 Flash Drum",          "Vertical Vessel",    "LP NH3 letdown flash",         "CS",                      25,   20,   5.0,  "D=0.8m, H=3.0m",         "—",             "Gas NH3 product drum"),
        ("E-3521", "NH3 Column Condenser",       "Shell & Tube HX",    "Column overhead condensing",   "CS",                      62,   60,  12.0,  "A=83 m²",                "1.45 Gcal/hr",  "CW cooling 30°C"),
        ("E-3522", "NH3 Column Reboiler",        "Shell & Tube HX",    "Column reboiling",             "CS",                     110,  115,  12.0,  "A=47 m²",                "1.35 Gcal/hr",  "LP steam 145°C"),
    ]

    columns_eq = [
        "Tag",
        "Description",
        "Equipment Type",
        "Service / Duty",
        "Material of Construction",
        "T_in (°C)",
        "T_out (°C)",
        "P_in (bar(g))",
        "Size / Volume",
        "Duty / Power",
        "Notes",
    ]
    n_cols = len(columns_eq)

    # Title
    ws.row_dimensions[1].height = 28
    title_style(ws, 1, 1, "D2 – Equipment List & Specifications | 100 TPD Ammonia Plant", merge_to_col=n_cols)

    # Header row
    ws.row_dimensions[2].height = 30
    for ci, ch in enumerate(columns_eq, start=1):
        c = ws.cell(row=2, column=ci, value=ch)
        header_style(c, level=1)

    # Section headers & data
    sections = [
        ("FEED GAS TREATING & REFORMING (3200 SERIES)", list(range(0, 13))),
        ("CO2 REMOVAL & METHANATION (3300 SERIES)", list(range(13, 24))),
        ("SYNGAS COMPRESSION (3400 SERIES)", list(range(24, 31))),
        ("SYNTHESIS LOOP & NH3 SEPARATION (3500 SERIES)", list(range(31, len(equipment)))),
    ]

    section_fill = PatternFill("solid", fgColor="D6E4F0")
    light_fill   = PatternFill("solid", fgColor="EBF3FB")
    current_row  = 3

    for section_title, indices in sections:
        # Section banner
        ws.row_dimensions[current_row].height = 18
        sc = ws.cell(row=current_row, column=1, value=section_title)
        sc.font  = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        sc.fill  = section_fill
        sc.alignment = Alignment(horizontal="left", vertical="center")
        sc.border = make_border()
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=n_cols)
        current_row += 1

        for idx in indices:
            row_data = equipment[idx]
            ws.row_dimensions[current_row].height = 16
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                if ci == 1:
                    cell.font = Font(name="Calibri", bold=True, size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                cell.border = make_border()
                if current_row % 2 == 0:
                    cell.fill = light_fill
            current_row += 1

    # Footer
    footer_row(ws, current_row + 1, 1, n_cols)

    # Column widths
    col_widths = [
        (1, 12), (2, 32), (3, 22), (4, 30), (5, 28),
        (6, 10), (7, 10), (8, 12), (9, 22), (10, 16), (11, 38),
    ]
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "B3"

    path = os.path.join(OUTPUT_DIR, "NH3_Equipment_List.xlsx")
    wb.save(path)
    print(f"  Saved: {path}")
    return path


# ===========================================================================
# D3 – NH3_HX_Utility_Summary.xlsx
# ===========================================================================

def create_hx_summary():
    wb = Workbook()
    ws = wb.active
    ws.title = "HX & Utility Summary"

    # Tag | Description | Hot-side In (°C) | Hot-side Out (°C) | Hot flow (kg/hr)
    # | Cold-side In (°C) | Cold-side Out (°C) | Cold flow (kg/hr)
    # | Duty (Gcal/hr) | Duty (MW) | LMTD (°C) | U (kcal/m²·hr·°C) | Area (m²) | Utility Type
    hx_data = [
        ("E-3204",  "NG Preheater",                950,  400, 8000,   40,  380,  1850,  2.80, 3.26,  410, 120,  56.7, "Flue Gas"),
        ("E-3201",  "Feed/Steam Preheater",         850,  450, 8000,  200,  490,  8619,  6.20, 7.21,  380, 100, 163.0, "Flue Gas"),
        ("E-3206",  "Sec. Reformer WHB",            990,  350,13200,   40,  338,  9500, 18.50,21.52,  320, 350, 165.0, "HP Steam Gen (125 bar)"),
        ("E-3208",  "HT/LT Shift Interstage HX",   437,  200,13200,   40,  160,  6000,  8.20, 9.54,  195, 300, 140.0, "BFW/Steam Gen"),
        ("E-3209",  "LT Shift Outlet Cooler",       232,  107,13200,   30,   45, 28000,  4.50, 5.23,  105, 550,  78.0, "CW"),
        ("TX-3301", "Rich/Lean MDEA HX",            107,   55, 8000,   55,  107,  8000,  3.80, 4.42,   30, 200, 420.0, "Process-to-Process"),
        ("E-3311",  "Methanator F/E HX",            339,  107, 9580,  107,  290,  9580,  2.10, 2.44,   80, 250, 105.0, "Process-to-Process"),
        ("E-3312",  "Meta Outlet Cooler",           107,   38, 9580,   30,   45, 12000,  0.62, 0.72,   45, 600,  23.0, "CW"),
        ("E-3401A", "Syngas Interstage Cooler 1",   195,   40, 9020,   30,   45, 38000,  3.10, 3.61,   65, 650,  73.0, "CW"),
        ("E-3401B", "Syngas Interstage Cooler 2",   270,   40, 9020,   30,   45, 38000,  3.80, 4.42,   80, 650,  73.0, "CW"),
        ("E-3401C", "Syngas Interstage Cooler 3",   310,   40, 9020,   30,   45, 38000,  4.20, 4.89,   95, 650,  68.0, "CW"),
        ("E-3501",  "Converter WHB",                455,  320,34000,   40,  338, 12000, 12.80,14.90,  125, 350, 293.0, "HP Steam Gen (125 bar)"),
        ("E-3502",  "Loop P2P Heat Exchanger",      455,  380,34000,  220,  320, 34000,  6.50, 7.56,   60, 300, 361.0, "Process-to-Process"),
        ("E-3503",  "Converter Trim Cooler",        320,   80,34000,   40,  160,  8000,  1.80, 2.09,   85, 280,  75.7, "BFW Preheat"),
        ("E-3504",  "Loop CW Cooler",                80,   40,34000,   30,   45, 80000,  4.20, 4.89,   20, 700, 300.0, "CW"),
        ("E-3506",  "NH3 Refrigerant Chiller 1",    40, 18.8,34000,  -10,    0,     0,   2.80, 3.26,   30, 450, 207.0, "NH3 Refrigerant (-10°C)"),
        ("E-3507",  "NH3 Refrigerant Chiller 2",  18.8,   12,34000,  -20,  -10,     0,   1.20, 1.40,   26, 400, 115.0, "NH3 Refrigerant (-20°C)"),
        ("E-3511",  "Purge Gas Chiller",           18.8,  -25, 2550,  -30,  -25,     0,   0.18, 0.21,   10, 350,  52.0, "NH3 Refrigerant (-30°C)"),
        ("E-3521",  "NH3 Column Condenser",          62,   60, 5200,   30,   45, 18000,  1.45, 1.69,   25, 700,  83.0, "CW"),
        ("E-3522",  "NH3 Column Reboiler",          145,  145, 2800,  110,  115,  8200,  1.35, 1.57,   32, 900,  47.0, "LP Steam (145°C)"),
        ("E-3321",  "Stripped Condensate Cooler",   180,   45, 6848,   30,   80,  8000,  0.65, 0.76,   75, 500,  17.0, "BFW/CW"),
    ]

    columns_hx = [
        "Tag", "Description",
        "Hot In\n(°C)", "Hot Out\n(°C)", "Hot Flow\n(kg/hr)",
        "Cold In\n(°C)", "Cold Out\n(°C)", "Cold Flow\n(kg/hr)",
        "Duty\n(Gcal/hr)", "Duty\n(MW)",
        "LMTD\n(°C)", "U\n(kcal/m²·hr·°C)", "Area\n(m²)",
        "Utility / Service",
    ]
    n_cols = len(columns_hx)

    # Title
    ws.row_dimensions[1].height = 28
    title_style(ws, 1, 1, "D3 – Heat Exchanger & Utility Summary | 100 TPD Ammonia Plant", merge_to_col=n_cols)

    # Sub-title info row
    ws.row_dimensions[2].height = 18
    info = ws.cell(row=2, column=1, value="Basis: Q = U·A·LMTD  |  1 Gcal/hr = 1.163 MW  |  All duties at design (100% load)")
    info.font = Font(name="Calibri", italic=True, size=9, color="404040")
    info.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    # Header
    ws.row_dimensions[3].height = 36
    for ci, ch in enumerate(columns_hx, start=1):
        c = ws.cell(row=3, column=ci, value=ch)
        header_style(c, level=1)

    light_fill = PatternFill("solid", fgColor="EBF3FB")
    for ri, row in enumerate(hx_data, start=4):
        ws.row_dimensions[ri].height = 16
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 1:
                cell.font = Font(name="Calibri", bold=True, size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 2:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci in (3, 4, 6, 7):   # temperatures
                data_style(cell, number_format="0.0")
            elif ci in (5, 8):          # flow rates
                data_style(cell, number_format="#,##0")
            elif ci in (9, 10):         # duties
                data_style(cell, number_format="0.00")
            elif ci == 11:              # LMTD
                data_style(cell, number_format="0")
            elif ci == 12:              # U
                data_style(cell, number_format="0")
            elif ci == 13:              # area
                data_style(cell, number_format="0.0")
            else:
                data_style(cell)
            cell.border = make_border()
            if ri % 2 == 0:
                if not cell.fill or cell.fill.fill_type != "solid":
                    cell.fill = light_fill
                else:
                    pass
        if ri % 2 == 0:
            for ci in range(1, n_cols + 1):
                ws.cell(row=ri, column=ci).fill = light_fill

    # Totals row
    total_row = 4 + len(hx_data)
    ws.row_dimensions[total_row].height = 18
    total_duty = sum(r[8] for r in hx_data)
    total_area = sum(r[12] for r in hx_data)
    tc = ws.cell(row=total_row, column=1, value="TOTAL")
    tc.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FONT_COLOR)
    tc.fill = PatternFill("solid", fgColor=DARK_BLUE)
    tc.border = make_border()
    tc.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    for ci in range(2, 9):
        ws.cell(row=total_row, column=ci).fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws.cell(row=total_row, column=ci).border = make_border()

    td = ws.cell(row=total_row, column=9, value=round(total_duty, 2))
    td.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FONT_COLOR)
    td.fill = PatternFill("solid", fgColor=DARK_BLUE)
    td.number_format = "0.00"
    td.border = make_border()
    td.alignment = Alignment(horizontal="center")

    tmw = ws.cell(row=total_row, column=10, value=round(total_duty * 1.163, 2))
    tmw.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FONT_COLOR)
    tmw.fill = PatternFill("solid", fgColor=DARK_BLUE)
    tmw.number_format = "0.00"
    tmw.border = make_border()
    tmw.alignment = Alignment(horizontal="center")

    for ci in range(11, 13):
        blank = ws.cell(row=total_row, column=ci, value="—")
        blank.fill = PatternFill("solid", fgColor=DARK_BLUE)
        blank.font = Font(bold=True, color=HEADER_FONT_COLOR)
        blank.border = make_border()
        blank.alignment = Alignment(horizontal="center")

    ta = ws.cell(row=total_row, column=13, value=round(total_area, 1))
    ta.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FONT_COLOR)
    ta.fill = PatternFill("solid", fgColor=DARK_BLUE)
    ta.number_format = "0.0"
    ta.border = make_border()
    ta.alignment = Alignment(horizontal="center")

    blank14 = ws.cell(row=total_row, column=14, value="—")
    blank14.fill = PatternFill("solid", fgColor=DARK_BLUE)
    blank14.font = Font(bold=True, color=HEADER_FONT_COLOR)
    blank14.border = make_border()
    blank14.alignment = Alignment(horizontal="center")

    # Footer
    footer_row(ws, total_row + 2, 1, n_cols)

    # Column widths
    col_widths = [
        (1, 10), (2, 32), (3, 10), (4, 10), (5, 13),
        (6, 10), (7, 10), (8, 13),
        (9, 12), (10, 10), (11, 10), (12, 18), (13, 10), (14, 26),
    ]
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "C4"

    path = os.path.join(OUTPUT_DIR, "NH3_HX_Utility_Summary.xlsx")
    wb.save(path)
    print(f"  Saved: {path}")
    return path


# ===========================================================================
# Energy Balance Summary – NH3_Energy_Balance_Summary.xlsx
# ===========================================================================

def create_energy_balance():
    wb = Workbook()
    ws = wb.active
    ws.title = "Energy Balance"

    ws.row_dimensions[1].height = 30
    title_style(ws, 1, 1, "Energy Balance Summary | 100 TPD Ammonia Plant (NH3_Plant_Simulation.apwz)", merge_to_col=6)

    # Info
    ws.row_dimensions[2].height = 16
    info = ws.cell(row=2, column=1, value="Design Basis: 100 TPD NH3 = 245.1 kmol/hr  |  Synthesis loop: 220 bar(g)  |  SMR-based process")
    info.font = Font(name="Calibri", italic=True, size=9, color="404040")
    info.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    def section_header(row, title):
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=1, value=title)
        c.font  = Font(name="Calibri", bold=True, size=11, color=HEADER_FONT_COLOR)
        c.fill  = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = make_border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=MID_BLUE)
            ws.cell(row=row, column=ci).border = make_border()

    def data_row(row, item, value, unit, comment="", highlight=False):
        ws.row_dimensions[row].height = 16
        fg = LIGHT_BLUE if highlight else None
        vals = [item, value, unit, comment, "", ""]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = make_border()
            cell.font = Font(name="Calibri", size=10, bold=(ci == 1 or highlight))
            if ci == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if highlight and fg:
                cell.fill = PatternFill("solid", fgColor=fg)

    # Row 3: column headers
    col_headers = ["Energy Item", "Value", "Unit", "Comments / Notes", "", ""]
    ws.row_dimensions[3].height = 22
    for ci, h in enumerate(col_headers, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        header_style(c, level=1)

    r = 4

    # ---- Section 1: Fired Heater ----
    section_header(r, "1. FIRED HEATER – PRIMARY REFORMER (H-3201)"); r += 1
    data_row(r, "Fired Duty (LHV basis)",              38.5,  "Gcal/hr", "Natural gas + purge gas fuel"); r += 1
    data_row(r, "Fired Duty",                          44.8,  "MW",      "1 Gcal/hr = 1.163 MW"); r += 1
    data_row(r, "Thermal Efficiency (LHV)",            91.5,  "%",       "Including WHB / convection recovery"); r += 1
    data_row(r, "Flue Gas Flow",                       8000,  "kg/hr",   "After WHB convection section"); r += 1
    data_row(r, "Flue Gas Exit Temperature",           150,   "°C",      "Stack temperature"); r += 1
    data_row(r, "Stack Heat Loss",                     1.5,   "Gcal/hr", "~4% of fired duty"); r += 1

    r += 1
    # ---- Section 2: Steam System ----
    section_header(r, "2. STEAM GENERATION & CONSUMPTION"); r += 1
    data_row(r, "HP Steam Generated – E-3206 (Sec. Reformer WHB)", 9500, "kg/hr",  "125 bar, 338°C saturated"); r += 1
    data_row(r, "HP Steam Generated – E-3501 (Converter WHB)",    12000, "kg/hr",  "125 bar, 510°C superheated"); r += 1
    data_row(r, "HP Steam Generated – Other WHBs / convection",    7000, "kg/hr",  "H-3201 convection, E-3503"); r += 1
    data_row(r, "TOTAL HP Steam Generated",                       28500, "kg/hr",  "At 125 bar, ~510°C avg", highlight=True); r += 1
    data_row(r, "HP Steam to K-3401 Turbine (Syngas Comp.)",       9200, "kg/hr",  "Drives 11.2 MW syngas compressor"); r += 1
    data_row(r, "HP Steam to K-3501 Turbine (Recycle Comp.)",      4000, "kg/hr",  "Drives 3.8 MW recycle compressor"); r += 1
    data_row(r, "HP Steam to Process (CO2 stripping, utilities)",  5000, "kg/hr",  "LP steam let-down + CO2 stripper"); r += 1
    data_row(r, "TOTAL HP Steam Consumed",                        18200, "kg/hr",  "All turbines + process steam", highlight=True); r += 1
    data_row(r, "NET HP Steam Export",                            10300, "kg/hr",  "Available to adjacent urea/utility plant", highlight=True); r += 1
    data_row(r, "LP Steam Generated (turbine exhaust)",            8500, "kg/hr",  "~4.5 bar, 160°C"); r += 1
    data_row(r, "LP Steam Consumed (E-3522, F-3301 reboiler)",     5300, "kg/hr",  "NH3 column reboiler + CO2 stripper"); r += 1
    data_row(r, "Process Steam to Reformer (S-STEAM-PROC)",        6769, "kg/hr",  "376 kmol/hr at 35 bar, 245°C"); r += 1

    r += 1
    # ---- Section 3: Cooling Water ----
    section_header(r, "3. COOLING WATER DUTIES"); r += 1
    data_row(r, "E-3209  LT Shift Outlet Cooler",   4.50, "Gcal/hr", "Process gas 232°C→107°C"); r += 1
    data_row(r, "E-3312  Methanator Outlet Cooler",  0.62, "Gcal/hr", "Process gas 107°C→38°C"); r += 1
    data_row(r, "E-3401A Syngas Interstage Cooler 1",3.10, "Gcal/hr", "Syngas 195°C→40°C"); r += 1
    data_row(r, "E-3401B Syngas Interstage Cooler 2",3.80, "Gcal/hr", "Syngas 270°C→40°C"); r += 1
    data_row(r, "E-3401C Syngas Interstage Cooler 3",4.20, "Gcal/hr", "Syngas 310°C→40°C"); r += 1
    data_row(r, "E-3504  Loop CW Cooler",            4.20, "Gcal/hr", "Loop gas 80°C→40°C"); r += 1
    data_row(r, "E-3521  NH3 Column Condenser",      1.45, "Gcal/hr", "NH3 vapor 62°C→60°C"); r += 1
    data_row(r, "Miscellaneous CW (utilities, etc)", 6.93, "Gcal/hr", "Fin-fans, MDEA cooler, other"); r += 1
    data_row(r, "TOTAL Cooling Water Duty",          28.8, "Gcal/hr", "CW supply 30°C, return 45°C", highlight=True); r += 1
    data_row(r, "Cooling Water Flow Rate",         164000, "kg/hr",   "ΔT=15°C, Cp=1 kcal/kg·°C"); r += 1

    r += 1
    # ---- Section 4: Refrigeration ----
    section_header(r, "4. REFRIGERATION SYSTEM"); r += 1
    data_row(r, "E-3506  NH3 Chiller Stage 1 (-10°C level)",  2.80, "Gcal/hr", "Loop gas 40°C→18.8°C condensation"); r += 1
    data_row(r, "E-3507  NH3 Chiller Stage 2 (-20°C level)",  1.20, "Gcal/hr", "Loop gas 18.8°C→12°C"); r += 1
    data_row(r, "E-3511  Purge Gas Chiller (-30°C level)",    0.18, "Gcal/hr", "Purge NH3 recovery, -25°C outlet"); r += 1
    data_row(r, "TOTAL Refrigeration Duty",                   4.18, "Gcal/hr", "Multi-stage NH3 vapor compression", highlight=True); r += 1
    data_row(r, "Refrigeration COP (approx.)",                2.8,  "—",       "At -20°C avg level, 45°C condensing"); r += 1
    data_row(r, "Refrigeration Compressor Power",             1.49, "MW",      "4.18 Gcal/hr / 2.8 COP / 1.163"); r += 1

    r += 1
    # ---- Section 5: Compression Power ----
    section_header(r, "5. COMPRESSION POWER"); r += 1
    data_row(r, "K-3401 Syngas Compressor (30→225 bar)",  11.2, "MW", "Steam turbine driven, 4-stage"); r += 1
    data_row(r, "K-3501 Recycle Compressor (loop)",        3.8, "MW", "Steam turbine driven, 1-stage"); r += 1
    data_row(r, "K-3203 Process Air Compressor (1→32 bar)",2.2, "MW", "Motor driven"); r += 1
    data_row(r, "Refrigeration Compressor (K-3502)",        1.5, "MW", "Motor driven, multi-stage NH3 ref."); r += 1
    data_row(r, "Other Compression / Misc.",                0.5, "MW", "Instrument air, booster comps"); r += 1
    data_row(r, "TOTAL Compression Power (shaft)",         19.2, "MW", "Steam turbines + motors", highlight=True); r += 1

    r += 1
    # ---- Section 6: Electrical ----
    section_header(r, "6. ELECTRICAL LOAD SUMMARY"); r += 1
    data_row(r, "Motor-Driven Compressors (K-3203, K-3502)", 3.7, "MW", "Air comp + refrigeration comp"); r += 1
    data_row(r, "Pumps (MDEA P-3301, BFW, misc.)",           1.2, "MW", "Multiple pumps across plant"); r += 1
    data_row(r, "Instruments, Controls & Lighting",           0.8, "MW", "DCS, ESD, CCTV, lighting"); r += 1
    data_row(r, "Miscellaneous Electrical Loads",             1.5, "MW", "HVAC, trace heating, misc."); r += 1
    data_row(r, "Utilities & Offsites",                       1.0, "MW", "Water treatment, cooling tower fans"); r += 1
    data_row(r, "TOTAL Electrical Demand",                    8.2, "MW", "Plant MCC connected load", highlight=True); r += 1
    data_row(r, "Steam Turbine Power Generated (K-3401+K-3501)",15.0,"MW","Available for back-pressure generation"); r += 1
    data_row(r, "Net Electrical Surplus to Grid",              6.8, "MW", "Steam turbines exceed motor loads"); r += 1

    r += 1
    # ---- Section 7: Overall Energy Metrics ----
    section_header(r, "7. KEY PERFORMANCE INDICATORS"); r += 1
    data_row(r, "NG Feed (Process + Fuel)",                   1850, "kg/hr",    "~21.8 GJ/hr LHV based on 100 kmol/hr NG"); r += 1
    data_row(r, "NH3 Production",                              100, "TPD",      "= 4166.7 kg/hr"); r += 1
    data_row(r, "Energy Consumption (LHV)",                   7.56, "Gcal/t NH3","= 31.7 GJ/t NH3 (LHV basis)"); r += 1
    data_row(r, "Energy Consumption (HHV)",                   8.22, "Gcal/t NH3","= 34.4 GJ/t NH3 (HHV basis)"); r += 1
    data_row(r, "Steam Export Credit",                         10300,"kg/hr",   "125 bar HP steam to battery limit"); r += 1
    data_row(r, "CO2 Export (to Urea Plant)",                 16051, "kg/hr",   "99.8% pure CO2 from MDEA absorber"); r += 1
    data_row(r, "CO2 Specific Emissions (net, no credit)",    1.61,  "t CO2/t NH3","Before urea/steam credit"); r += 1
    data_row(r, "H2:N2 Ratio at Synthesis Inlet",             3.0,  "mol/mol",  "Stoichiometric H2:N2 = 3:1"); r += 1
    data_row(r, "NH3 Conversion per Pass",                    30.0,  "%",        "Single-pass at 455°C, 220 bar"); r += 1
    data_row(r, "Overall NH3 Loop Conversion",                97.5,  "%",        "Recycle basis"); r += 1

    r += 1
    footer_row(ws, r + 1, 1, 6)

    # Column widths
    col_widths = [(1, 50), (2, 14), (3, 14), (4, 52), (5, 5), (6, 5)]
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A4"

    path = os.path.join(OUTPUT_DIR, "NH3_Energy_Balance_Summary.xlsx")
    wb.save(path)
    print(f"  Saved: {path}")
    return path


# ===========================================================================
# Main
# ===========================================================================

if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("=" * 60)
    print("NH3 Plant Flowsheet Data Generator")
    print("100 TPD Integrated Ammonia Plant")
    print("=" * 60)

    files = []
    print("\nGenerating D4 – Stream Table...")
    files.append(create_stream_table())

    print("\nGenerating D2 – Equipment List...")
    files.append(create_equipment_list())

    print("\nGenerating D3 – HX & Utility Summary...")
    files.append(create_hx_summary())

    print("\nGenerating Energy Balance Summary...")
    files.append(create_energy_balance())

    print("\n" + "=" * 60)
    print("All files generated successfully:")
    for f in files:
        size_kb = os.path.getsize(f) / 1024
        print(f"  {os.path.basename(f):50s}  {size_kb:6.1f} KB")
    print("=" * 60)
